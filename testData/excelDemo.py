import openpyxl

book = openpyxl.load_workbook("C:\\Users\\gulna\\Downloads\\TestData.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=3)
print(cell.value)

Dict = {}

sheet.cell(row=2, column=2).value = "Gulnawaz"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

firstName = sheet['B2'].value
lastName = sheet['C4'].value
print(firstName + " " + lastName)

for rowValues in range(1, sheet.max_row+1):
    if sheet.cell(row=rowValues, column=1).value == "TestCase1":
        for columnValues in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=columnValues).value] = sheet.cell(row=rowValues, column=columnValues).value
            # print(sheet.cell(row=rowValues, column=columnValues).value)
print(Dict)
