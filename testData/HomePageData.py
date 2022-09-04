import openpyxl


class HomePageData:

    # test_HomePage_data = [{"name": "Gulnawaz",
    #                        "email": "gulnawaz@gmail.com",
    #                        "password": "Password",
    #                        "gender": "Female"},
    #                       {"name": "Dilnawaz",
    #                        "email": "dilnawaz@gmail.com",
    #                        "password": "Password",
    #                        "gender": "Male"}]

    @staticmethod
    def getTestData(testCaseName):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\gulna\\Downloads\\TestData.xlsx")
        sheet = book.active
        for rowValues in range(1, sheet.max_row + 1):
            if sheet.cell(row=rowValues, column=1).value == testCaseName:
                for columnValues in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=columnValues).value] = sheet.cell(row=rowValues, column=columnValues).value
                    # print(sheet.cell(row=rowValues, column=columnValues).value)
        return [Dict]

